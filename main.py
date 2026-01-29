from flask import Flask, request, make_response
import requests
import os
import time
import threading
import re
import hashlib
import sqlite3                     # === ДОБАВЛЕНО ===
from openpyxl import load_workbook

app = Flask(__name__)

# =========================================================
# ======================= ENV ==============================
# =========================================================

BOT_TOKEN = os.environ["BOT_TOKEN"]
ASANA_TOKEN = os.environ["ASANA_TOKEN"]
ASANA_PROJECT_ID = os.environ["ASANA_PROJECT_ID"]
ASANA_ASSIGNEE_ID = os.environ["ASANA_ASSIGNEE_ID"]

TELEGRAM_API = f"https://api.telegram.org/bot{BOT_TOKEN}"
ASANA_HEADERS = {"Authorization": f"Bearer {ASANA_TOKEN}"}

REQUIRED_PHOTOS = 3
SLA_SECONDS = 30 * 60                 # НЕ УДАЛЯЕМ
REWARDS_FILE = "data/rewards.xlsx"

PHOTO_DB = "data/photo_hashes.db"     # === ДОБАВЛЕНО ===

# =========================================================
# ======================= STATE =============================
# =========================================================

user_states = {}
user_data = {}
sent_notifications = set()

# =========================================================
# ======================= SQLITE ============================
# =========================================================

def init_db():
    conn = sqlite3.connect(PHOTO_DB)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS photo_hashes (
            hash TEXT PRIMARY KEY
        )
    """)
    conn.commit()
    conn.close()

def photo_hash_exists(h):
    conn = sqlite3.connect(PHOTO_DB)
    c = conn.cursor()
    c.execute("SELECT 1 FROM photo_hashes WHERE hash = ?", (h,))
    exists = c.fetchone() is not None
    conn.close()
    return exists

def save_photo_hash(h):
    conn = sqlite3.connect(PHOTO_DB)
    c = conn.cursor()
    try:
        c.execute("INSERT INTO photo_hashes(hash) VALUES (?)", (h,))
        conn.commit()
    except sqlite3.IntegrityError:
        pass
    conn.close()

init_db()

# =========================================================
# ======================= TEXTS =============================
# =========================================================

PROCESS_TEXT = {
    "ru": "⏳ Ваша заявка подана и будет обработана в течение 3 рабочих дней.",
    "uz": "⏳ Arizangiz qabul qilindi va 3 ish kuni ichida ko‘rib chiqiladi."
}

# ⬇️ TEXTS — БЕЗ СОКРАЩЕНИЙ, как в твоём исходнике
TEXTS = {
    "ru": {
        "choose_lang": "🌐 Выберите язык",
        "menu": "Выберите нужный раздел:",
        "menu_buttons": ["📸 Фото-контроль", "🎁 Вознаграждение", "📄 Статус заявки"],
        "start_info": (
            "🚗 *Фото-контроль брендированного автомобиля*\n\n"
            "1️⃣ Введите ФИО\n"
            "2️⃣ Введите табельный номер\n"
            "3️⃣ Отправьте 3 фото автомобиля\n\n"
            "⚠️ Брендинг и госномер должны быть чётко видны."
        ),
        "fio": "✍️ *Шаг 1 из 3*\nВведите ФИО полностью",
        "tab": (
            "🔢 *Шаг 2 из 3*\n"
            "Введите табельный номер\n\n"
            "📌 Пример: `12345`"
        ),
        "tab_invalid": "❌ Табельный номер должен состоять из *5 цифр*.",
        "photo": "📸 *Шаг 3 из 3*\nОтправьте 3 фото по одному.",
        "photo_duplicate": "❌ Это фото уже использовалось ранее.",
        "photo_wrong_state": "❌ Сейчас нельзя отправлять фото.",
        "photo_done": "✅ Все фото получены.\nНажмите «Завершить».",
        "submitted": "⏳ Ваша заявка подана и будет обработана в течение 3 рабочих дней.",
        "approved": "✅ *Фото-контроль пройден*",
        "rejected": "❌ *Фото-контроль не пройден*\nПричина:\n{reason}",
        "need_photos": "❌ Нужно отправить ровно 3 фото.",
        "reward_not_allowed": "🎁 Вознаграждение доступно только после успешного фото-контроля.",
        "reward_not_found": "🎁 Данные по вознаграждению не найдены.",
        "reward_info": (
            "🎁 *Вознаграждение*\n\n"
            "👤 {fio}\n"
            "📅 Отработано дней: {days}\n"
            "💰 Сумма: {amount}"
        ),
        "copy_code": "📋 Скопировать промокод",
        "status_no_task": "📄 У вас нет активной заявки.",
        "status_text": (
            "📄 *Статус заявки*\n\n"
            "🆔 ID: {gid}\n"
            "⏳ Статус: {status}"
        ),
        "status_map": {
            "pending": "На проверке",
            "approved": "Одобрено",
            "rejected": "Отклонено"
        },
        "buttons": {
            "start": "▶️ Начать",
            "cancel": "❌ Отменить",
            "finish": "✅ Завершить"
        }
    },

    "uz": {
        "choose_lang": "🌐 Tilni tanlang",
        "menu": "Kerakli bo‘limni tanlang:",
        "menu_buttons": ["📸 Foto-nazorat", "🎁 Mukofot", "📄 Ariza holati"],
        "start_info": (
            "🚗 *Avtomobil foto-nazorati*\n\n"
            "1️⃣ F.I.Sh kiriting\n"
            "2️⃣ Tabel raqamini kiriting\n"
            "3️⃣ 3 ta foto yuboring"
        ),
        "fio": "✍️ *1-bosqich*\nF.I.Sh kiriting",
        "tab": "🔢 *2-bosqich*\n📌 Misol: `12345`",
        "tab_invalid": "❌ Tabel raqami 5 ta raqamdan iborat.",
        "photo": "📸 *3-bosqich*\n3 ta foto yuboring.",
        "photo_duplicate": "❌ Bu rasm avval ishlatilgan.",
        "photo_wrong_state": "❌ Hozir rasm yuborib bo‘lmaydi.",
        "photo_done": "✅ Barcha foto qabul qilindi.\n«Yakunlash» ni bosing.",
        "submitted": "⏳ Arizangiz qabul qilindi va 3 ish kuni ichida ko‘rib chiqiladi.",
        "approved": "✅ Foto-nazoratdan o‘tildi",
        "rejected": "❌ O‘tilmadi\nSabab:\n{reason}",
        "need_photos": "❌ Aniq 3 ta foto kerak.",
        "reward_not_allowed": "🎁 Mukofot faqat tasdiqlangandan so‘ng beriladi.",
        "reward_not_found": "🎁 Mukofot topilmadi.",
        "reward_info": (
            "🎁 *Mukofot*\n\n"
            "👤 {fio}\n"
            "📅 Ishlangan kunlar: {days}\n"
            "💰 Summa: {amount}"
        ),
        "copy_code": "📋 Promokodni nusxalash",
        "status_no_task": "📄 Sizda faol ariza yo‘q.",
        "status_text": (
            "📄 *Ariza holati*\n\n"
            "🆔 ID: {gid}\n"
            "⏳ Holat: {status}"
        ),
        "status_map": {
            "pending": "Tekshiruvda",
            "approved": "Tasdiqlandi",
            "rejected": "Rad etildi"
        },
        "buttons": {
            "start": "▶️ Boshlash",
            "cancel": "❌ Bekor qilish",
            "finish": "✅ Yakunlash"
        }
    }
}

# =========================================================
# ======================= HELPERS ===========================
# =========================================================

def kb(buttons):
    return {"keyboard": [[{"text": b}] for b in buttons], "resize_keyboard": True}

def inline_kb(text, data):
    return {"inline_keyboard": [[{"text": text, "callback_data": data}]]}

def send(chat_id, text, keyboard=None):
    payload = {"chat_id": chat_id, "text": text, "parse_mode": "Markdown"}
    if keyboard:
        payload["reply_markup"] = keyboard
    requests.post(f"{TELEGRAM_API}/sendMessage", json=payload, timeout=10)

def download_file(file_id):
    info = requests.get(f"{TELEGRAM_API}/getFile", params={"file_id": file_id}).json()
    path = info["result"]["file_path"]
    return requests.get(f"https://api.telegram.org/file/bot{BOT_TOKEN}/{path}").content

def get_asana_status(task_gid):
    r = requests.get(
        f"https://app.asana.com/api/1.0/tasks/{task_gid}",
        headers=ASANA_HEADERS,
        params={"opt_fields": "approval_status"}
    )
    if r.status_code != 200:
        return None
    return r.json()["data"]["approval_status"]

# =========================================================
# ======================= TELEGRAM ==========================
# =========================================================

@app.route("/webhook", methods=["POST"])
def telegram():
    data = request.json or {}

    if "callback_query" in data:
        cq = data["callback_query"]
        cid = cq["message"]["chat"]["id"]
        code = cq["data"].replace("COPY_", "")
        send(cid, code)
        return "ok"

    msg = data.get("message")
    if not msg:
        return "ok"

    cid = msg["chat"]["id"]
    txt = msg.get("text")
    photos = msg.get("photo")
    state = user_states.get(cid)
    lang = user_data.get(cid, {}).get("lang", "ru")
    btn = TEXTS[lang]["buttons"]

    # === ЖЁСТКИЙ БЛОКЕР ===
    if state == "WAIT_RESULT":
        send(cid, PROCESS_TEXT[lang])
        return "ok"

    if txt == "/start":
        user_states[cid] = "LANG"
        user_data[cid] = {}
        send(cid, TEXTS["ru"]["choose_lang"], kb(["Русский 🇷🇺", "O‘zbek 🇺🇿"]))
        return "ok"

    if state == "LANG":
        lang = "uz" if "O‘zbek" in txt else "ru"
        user_data[cid]["lang"] = lang
        user_states[cid] = "MENU"
        send(cid, TEXTS[lang]["menu"], kb(TEXTS[lang]["menu_buttons"]))
        return "ok"

    if txt in TEXTS[lang]["menu_buttons"]:
        if "Фото" in txt or "Foto" in txt:
            user_states[cid] = "READY"
            send(cid, TEXTS[lang]["start_info"], kb([btn["start"]]))
        else:
            task_gid = user_data.get(cid, {}).get("task_gid")
            if not task_gid or get_asana_status(task_gid) != "approved":
                send(cid, TEXTS[lang]["reward_not_allowed"])
                return "ok"

            reward = get_reward(cid)
            if not reward:
                send(cid, TEXTS[lang]["reward_not_found"])
                return "ok"

            fio, code, amount, days = reward
            send(
                cid,
                TEXTS[lang]["reward_info"].format(
                    fio=fio, amount=amount, days=days
                ),
                inline_kb(TEXTS[lang]["copy_code"], f"COPY_{code}")
            )
        return "ok"

    if state == "READY" and txt == btn["start"]:
        user_states[cid] = "WAIT_FIO"
        user_data[cid]["photos_count"] = 0
        send(cid, TEXTS[lang]["fio"])
        return "ok"

    if state == "WAIT_FIO":
        user_data[cid]["fio"] = txt
        user_states[cid] = "WAIT_TAB"
        send(cid, TEXTS[lang]["tab"])
        return "ok"

    if state == "WAIT_TAB":
        if not re.fullmatch(r"\d{5}", txt):
            send(cid, TEXTS[lang]["tab_invalid"])
            return "ok"
        user_data[cid]["tab"] = txt
        user_states[cid] = "WAIT_PHOTO"
        send(cid, TEXTS[lang]["photo"])
        return "ok"

    if state == "WAIT_PHOTO" and photos:
        file_bytes = download_file(photos[-1]["file_id"])
        file_hash = hashlib.sha256(file_bytes).hexdigest()

        if photo_hash_exists(file_hash):
            send(cid, TEXTS[lang]["photo_duplicate"])
            return "ok"

        save_photo_hash(file_hash)
        user_data[cid]["photos_count"] += 1

        if user_data[cid]["photos_count"] == REQUIRED_PHOTOS:
            send(cid, TEXTS[lang]["photo_done"], kb([btn["finish"]]))
        return "ok"

    if state == "WAIT_PHOTO" and txt == btn["finish"]:
        if user_data[cid]["photos_count"] != REQUIRED_PHOTOS:
            send(cid, TEXTS[lang]["need_photos"])
            return "ok"

        task_gid = create_asana_task(
            user_data[cid]["fio"],
            user_data[cid]["tab"],
            cid,
            lang
        )

        user_data[cid]["task_gid"] = task_gid
        user_states[cid] = "WAIT_RESULT"
        send(cid, TEXTS[lang]["submitted"], {"remove_keyboard": True})
        return "ok"

    return "ok"

# =========================================================
# ======================= REWARDS ===========================
# =========================================================

def get_reward(chat_id):
    if not os.path.exists(REWARDS_FILE):
        return None

    wb = load_workbook(REWARDS_FILE, data_only=True)
    ws = wb.active

    headers = {str(c.value).strip(): i for i, c in enumerate(ws[1])}

    for row in ws.iter_rows(min_row=2, values_only=True):
        tg_id = row[headers["Telegram ID"]]
        if tg_id and str(tg_id).strip() == str(chat_id):
            return (
                row[headers["ФИО"]],
                row[headers["Промокод"]],
                row[headers["Сумма"]],
                row[headers["Отработанные дни"]],
            )
    return None

# =========================================================
# ======================= ASANA =============================
# =========================================================

@app.route("/asana", methods=["GET", "POST"])
def asana():
    secret = request.headers.get("X-Hook-Secret")
    if secret:
        r = make_response("")
        r.headers["X-Hook-Secret"] = secret
        return r

    for e in (request.json or {}).get("events", []):
        gid = e.get("resource", {}).get("gid")
        if gid:
            threading.Thread(
                target=process_task,
                args=(gid,),
                daemon=True
            ).start()
    return "ok"

def process_task(task_gid):
    for _ in range(6):
        time.sleep(2)
        r = requests.get(
            f"https://app.asana.com/api/1.0/tasks/{task_gid}",
            headers=ASANA_HEADERS,
            params={"opt_fields": "approval_status,custom_fields.name,custom_fields.display_value"}
        )
        if r.status_code != 200:
            continue

        data = r.json()["data"]
        status = data["approval_status"]
        if status == "pending":
            continue

        key = f"{task_gid}:{status}"
        if key in sent_notifications:
            return
        sent_notifications.add(key)

        lang, reason = get_task_lang_and_comment(task_gid)
        text = TEXTS[lang]["approved"] if status == "approved" else TEXTS[lang]["rejected"].format(reason=reason)

        for f in data["custom_fields"]:
            if f["name"] == "Telegram ID":
                chat_id = int(f["display_value"])
                send(chat_id, text)
                user_states[chat_id] = "MENU"
                user_data[chat_id]["lang"] = lang
                send(chat_id, TEXTS[lang]["menu"], kb(TEXTS[lang]["menu_buttons"]))
        return

def get_task_lang_and_comment(task_gid):
    task = requests.get(
        f"https://app.asana.com/api/1.0/tasks/{task_gid}",
        headers=ASANA_HEADERS,
        params={"opt_fields": "notes"}
    ).json()["data"]

    lang = "uz" if "LANG:uz" in task.get("notes", "") else "ru"

    stories = requests.get(
        f"https://app.asana.com/api/1.0/tasks/{task_gid}/stories",
        headers=ASANA_HEADERS
    ).json()["data"]

    for s in reversed(stories):
        if s.get("type") == "comment":
            return lang, s.get("text")

    return lang, TEXTS[lang]["rejected"].format(reason="Причина не указана")

# =========================================================
# ======================= ROOT ==============================
# =========================================================

@app.route("/", methods=["GET", "HEAD"])
def root():
    return "", 200

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))


















